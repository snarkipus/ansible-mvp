"""Minimal derived report generation for the provenance MVP."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import struct
import tempfile
import zlib
from importlib import import_module
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from provenance.hashing import hash_artifact
from provenance.inventory import InventoryRecord, inventory_files, with_sha256
from provenance.paths import validate_run_id

REPORT_FILENAMES = ("summary.xlsx", "chart.png", "briefing.pptx")
RgbColor = tuple[int, int, int]


def build_report_products(
    *,
    run_id: str,
    workspace_root: Path | str = Path("."),
    required_csv: Path | str | None = None,
    ad_hoc_csv: Path | str | None = None,
) -> tuple[InventoryRecord, ...]:
    """Generate the MVP XLSX, PNG chart, and PPTX report products.

    Products are always written under
    ``runs/{run_id}/provenance/products/reports`` and summarized as hashed
    derived-product inventory records.
    """

    validate_run_id(run_id)

    root = Path(workspace_root).expanduser().resolve()
    provenance_root = root / "runs" / run_id / "provenance"
    extracted_root = provenance_root / "products" / "extracted"
    required_path = _resolve_product_path(required_csv, extracted_root / "required.csv")
    ad_hoc_path = _resolve_product_path(ad_hoc_csv, extracted_root / "ad_hoc.csv")
    reports_root = provenance_root / "products" / "reports"
    reports_root.mkdir(parents=True, exist_ok=True)

    required_bytes = _require_bound_report_input_validation(
        provenance_root, "required_extract", required_path
    )
    ad_hoc_bytes = _require_bound_report_input_validation(
        provenance_root, "ad_hoc_extract", ad_hoc_path
    )

    required_rows = _read_csv(required_bytes)
    ad_hoc_rows = _read_csv(ad_hoc_bytes)

    final_paths = {name: reports_root / name for name in REPORT_FILENAMES}
    temporary_paths = {
        name: _temporary_report_path(reports_root, Path(name).suffix) for name in REPORT_FILENAMES
    }
    try:
        _write_summary_workbook(temporary_paths["summary.xlsx"], required_rows, ad_hoc_rows)
        _write_chart(temporary_paths["chart.png"], ad_hoc_rows)
        _write_briefing(temporary_paths["briefing.pptx"], run_id, required_rows, ad_hoc_rows)
        _validate_report_products(temporary_paths)
        for name in REPORT_FILENAMES:
            temporary_paths[name].replace(final_paths[name])
    finally:
        for path in temporary_paths.values():
            path.unlink(missing_ok=True)

    records = inventory_files(reports_root)
    return tuple(
        _report_inventory_record(record, reports_root / record.relative_path) for record in records
    )


def build_report_product_evidence(
    *,
    run_id: str,
    workspace_root: Path | str = Path("."),
) -> tuple[dict[str, str | int | None], ...]:
    """Generate reports and return manifest-ready derived product evidence."""

    records = build_report_products(run_id=run_id, workspace_root=workspace_root)
    return tuple(_report_evidence(record) for record in records)


def _temporary_report_path(root: Path, suffix: str) -> Path:
    descriptor, name = tempfile.mkstemp(prefix=".report.", suffix=suffix, dir=root)
    os.close(descriptor)
    return Path(name)


def _require_bound_report_input_validation(
    provenance_root: Path, name: str, product_path: Path
) -> bytes:
    receipt_path = provenance_root / "validations" / f"{name}.json"
    if not receipt_path.is_file():
        raise ValueError(f"report generation requires passed validation evidence: {receipt_path}")
    loaded = json.loads(receipt_path.read_text(encoding="utf-8"))
    if not isinstance(loaded, dict) or loaded.get("status") != "pass":
        raise ValueError(f"report generation requires passed {name} validation: {receipt_path}")

    expected_path = product_path.relative_to(provenance_root).as_posix()
    if loaded.get("path") != expected_path:
        raise ValueError(
            f"report generation rejected {name} validation with product path "
            f"{loaded.get('path')!r}; expected {expected_path!r}: {receipt_path}"
        )
    product_bytes = product_path.read_bytes()
    actual_size = len(product_bytes)
    actual_sha256 = hashlib.sha256(product_bytes).hexdigest()
    if loaded.get("size_bytes") != actual_size or loaded.get("sha256") != actual_sha256:
        raise ValueError(
            f"report generation rejected stale {name} validation: current CSV size/SHA-256 "
            f"does not match {receipt_path}; revalidate {product_path} before building reports"
        )
    return product_bytes


def _validate_report_products(paths: dict[str, Path]) -> None:
    workbook = load_workbook(paths["summary.xlsx"], read_only=True)
    workbook.close()
    if paths["chart.png"].read_bytes()[:8] != b"\x89PNG\r\n\x1a\n":
        raise ValueError("generated chart is not a valid PNG")
    presentation_class = getattr(import_module("pptx"), "Presentation")
    presentation_class(paths["briefing.pptx"])


def _resolve_product_path(path: Path | str | None, default: Path) -> Path:
    candidate = default if path is None else Path(path)
    candidate = candidate.expanduser().resolve()
    if not candidate.is_file():
        raise FileNotFoundError(f"required report input does not exist: {candidate}")
    return candidate


def _read_csv(product_bytes: bytes) -> list[dict[str, str]]:
    return list(csv.DictReader(io.StringIO(product_bytes.decode("utf-8"), newline="")))


def _write_summary_workbook(
    path: Path, required_rows: list[dict[str, str]], ad_hoc_rows: list[dict[str, str]]
) -> None:
    workbook = Workbook()
    summary = workbook.active
    if summary is None:
        raise RuntimeError("new workbook did not create an active summary sheet")
    summary.title = "summary"
    summary.append(("metric", "value"))
    summary.append(("required_rows", len(required_rows)))
    summary.append(("ad_hoc_groups", len(ad_hoc_rows)))

    required_sheet = workbook.create_sheet("required_extract")
    _append_rows(required_sheet, required_rows)
    ad_hoc_sheet = workbook.create_sheet("ad_hoc_extract")
    _append_rows(ad_hoc_sheet, ad_hoc_rows)
    workbook.save(path)


def _append_rows(sheet: Worksheet, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def _write_chart(path: Path, ad_hoc_rows: list[dict[str, str]]) -> None:
    values = [_safe_int(row.get("total_bytes")) for row in ad_hoc_rows] or [0]
    path.write_bytes(_bar_chart_png(values))


def _write_briefing(
    path: Path, run_id: str, required_rows: list[dict[str, str]], ad_hoc_rows: list[dict[str, str]]
) -> None:
    presentation_factory: Any = import_module("pptx").Presentation
    inches: Any = import_module("pptx.util").Inches
    presentation = presentation_factory()
    slide = presentation.slides.add_slide(presentation.slide_layouts[5])
    slide.shapes.title.text = "Synthetic provenance report"
    text_box = slide.shapes.add_textbox(inches(0.8), inches(1.4), inches(8), inches(2.2))
    text_frame = text_box.text_frame
    text_frame.text = f"Run ID: {run_id}"
    for line in (
        f"Required extract rows: {len(required_rows)}",
        f"Ad hoc groups: {len(ad_hoc_rows)}",
        "Products are generated under provenance/products/reports.",
    ):
        paragraph = text_frame.add_paragraph()
        paragraph.text = line
    presentation.save(path)


def _safe_int(value: str | None) -> int:
    try:
        return int(value or "0")
    except ValueError:
        return 0


def _bar_chart_png(values: list[int]) -> bytes:
    width = 320
    height = 180
    margin = 24
    background: RgbColor = (255, 255, 255)
    axis: RgbColor = (80, 80, 80)
    bar: RgbColor = (79, 129, 189)
    pixels: list[list[RgbColor]] = [[background for _x in range(width)] for _y in range(height)]
    for x_coord in range(margin, width - margin):
        pixels[height - margin][x_coord] = axis
    for y_coord in range(margin, height - margin + 1):
        pixels[y_coord][margin] = axis

    max_value = max(values) or 1
    slot_width = max(1, (width - (2 * margin)) // len(values))
    for index, value in enumerate(values):
        bar_height = int((height - (2 * margin) - 1) * value / max_value)
        start_x = margin + (index * slot_width) + 4
        end_x = min(margin + ((index + 1) * slot_width) - 4, width - margin - 1)
        top_y = height - margin - bar_height
        for y_coord in range(top_y, height - margin):
            for x_coord in range(start_x, end_x + 1):
                pixels[y_coord][x_coord] = bar

    raw_rows = b"".join(b"\x00" + b"".join(bytes(pixel) for pixel in row) for row in pixels)
    return b"".join(
        (
            b"\x89PNG\r\n\x1a\n",
            _png_chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0)),
            _png_chunk(b"IDAT", zlib.compress(raw_rows)),
            _png_chunk(b"IEND", b""),
        )
    )


def _png_chunk(kind: bytes, data: bytes) -> bytes:
    checksum = zlib.crc32(kind + data) & 0xFFFFFFFF
    return struct.pack(">I", len(data)) + kind + data + struct.pack(">I", checksum)


def _report_inventory_record(record: InventoryRecord, path: Path) -> InventoryRecord:
    sha256 = hash_artifact(path, display_path=f"products/reports/{record.relative_path}").sha256
    return with_sha256(record, sha256 or "")


def _report_evidence(record: InventoryRecord) -> dict[str, str | int | None]:
    payload = record.to_dict()
    payload.update(
        {
            "relative_path": f"provenance/products/reports/{record.relative_path}",
            "area_type": "product",
            "product_area": "reports",
            "role": "report_product",
            "producing_stage": "build_reports",
        }
    )
    return payload
